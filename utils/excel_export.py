"""
Excel Export v2.1
==================
Layout:
  Page 1: Combined histogram + summary table
  Per image: Page A (images + histogram + stats), Page B (raw data)

Histogram: bar chart with normal curve overlay, bin range labels
under each bar (e.g. "10-20nm"), no legend box.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import numpy as np
from datetime import datetime
import os
import cv2
import tempfile
import math

C_DARK_BLUE = "1A2B4A"
C_MID_BLUE = "2E5FA3"
C_LIGHT_GRAY = "F2F4F7"
C_WHITE = "FFFFFF"
C_TEXT_DARK = "1A1A2E"

_thin = Side(style="thin", color="CCCCCC")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_center = Alignment(horizontal="center", vertical="center")
_left_indent = Alignment(horizontal="left", vertical="center", indent=1)


def _hdr_fill(color=C_DARK_BLUE):
    return PatternFill("solid", fgColor=color)

def _hdr_font(size=11, bold=True, color=C_WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _body_font(size=10, bold=False, color=C_TEXT_DARK):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _alt_fill(alt):
    return PatternFill("solid", fgColor="EEF3FF" if not alt else C_LIGHT_GRAY)

def _set_col_w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _choose_unit(px_per_um):
    """Pick clean units. Returns (area_unit, area_mult, diam_unit, diam_mult).
    Multipliers convert from um/um² to the chosen unit."""
    if px_per_um <= 0:
        return "px²", 1.0, "px", 1.0
    um_per_px = 1.0 / px_per_um
    nm_per_px = um_per_px * 1000.0
    if nm_per_px < 50:
        return "nm²", 1e6, "nm", 1000.0
    else:
        return "µm²", 1.0, "µm", 1.0


def _fmt_num(val):
    """Format a number cleanly — no tiny exponents."""
    if abs(val) < 0.01 and val != 0:
        return f"{val:.4g}"
    if abs(val) >= 1000:
        return f"{val:.0f}"
    if abs(val) >= 10:
        return f"{val:.1f}"
    if abs(val) >= 1:
        return f"{val:.2f}"
    return f"{val:.3f}"


def _save_image_temp(image_bgr, max_width=400):
    h, w = image_bgr.shape[:2]
    if w > max_width:
        scale = max_width / w
        image_bgr = cv2.resize(image_bgr, (max_width, int(h * scale)))
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    cv2.imwrite(tmp.name, image_bgr)
    tmp.close()
    return tmp.name


# ==================================================================
# Histogram with normal curve
# ==================================================================

def _build_histogram(grains, px_per_um):
    """Build histogram data. Returns (bin_labels, counts, edges, diam_unit, diam_mult, values)."""
    _, _, diam_unit, diam_mult = _choose_unit(px_per_um)

    if px_per_um > 0:
        values = np.array([g.equivalent_diameter_um * diam_mult for g in grains])
    else:
        values = np.array([g.equivalent_diameter_px for g in grains])

    if len(values) < 2:
        return [], [], [], diam_unit, diam_mult, values

    n_bins = min(max(int(math.sqrt(len(values))), 5), 25)
    counts, edges = np.histogram(values, bins=n_bins)

    # Clean bin labels: "10-20nm" style
    def fmt_edge(v):
        if abs(v) >= 100:
            return f"{v:.0f}"
        elif abs(v) >= 10:
            return f"{v:.1f}"
        elif abs(v) >= 1:
            return f"{v:.1f}"
        else:
            return f"{v:.2f}"

    labels = [f"{fmt_edge(edges[i])}-{fmt_edge(edges[i+1])}{diam_unit}"
              for i in range(len(edges) - 1)]

    return labels, counts.tolist(), edges.tolist(), diam_unit, diam_mult, values


def _write_histogram_chart(ws, grains, px_per_um, start_row, start_col=1):
    """Write histogram + normal curve combo chart."""
    labels, counts, edges, diam_unit, diam_mult, values = _build_histogram(grains, px_per_um)
    if not labels:
        return start_row

    n_bins = len(labels)
    data_row = start_row

    # Column layout: A=Bin Label, B=Count, C=Normal Curve
    hdr_c = start_col
    ws.cell(data_row, hdr_c, "Bin Range").font = _hdr_font(10)
    ws.cell(data_row, hdr_c).fill = _hdr_fill(C_MID_BLUE)
    ws.cell(data_row, hdr_c).alignment = _center

    ws.cell(data_row, hdr_c + 1, "Count").font = _hdr_font(10)
    ws.cell(data_row, hdr_c + 1).fill = _hdr_fill(C_MID_BLUE)
    ws.cell(data_row, hdr_c + 1).alignment = _center

    ws.cell(data_row, hdr_c + 2, "Normal Fit").font = _hdr_font(10)
    ws.cell(data_row, hdr_c + 2).fill = _hdr_fill(C_MID_BLUE)
    ws.cell(data_row, hdr_c + 2).alignment = _center

    # Compute normal curve values at bin centers
    mu = float(np.mean(values))
    sigma = float(np.std(values))
    bin_width = edges[1] - edges[0]
    n_total = len(values)

    for i in range(n_bins):
        r = data_row + 1 + i
        ws.cell(r, hdr_c, labels[i]).font = _body_font(9)
        ws.cell(r, hdr_c).border = _border
        ws.cell(r, hdr_c).alignment = _center

        ws.cell(r, hdr_c + 1, counts[i]).font = _body_font()
        ws.cell(r, hdr_c + 1).border = _border
        ws.cell(r, hdr_c + 1).alignment = _center

        # Normal curve value at bin center
        bin_center = (edges[i] + edges[i + 1]) / 2.0
        if sigma > 0:
            norm_val = ((1.0 / (sigma * math.sqrt(2 * math.pi))) *
                        math.exp(-0.5 * ((bin_center - mu) / sigma) ** 2))
            norm_val *= bin_width * n_total
        else:
            norm_val = 0
        ws.cell(r, hdr_c + 2, round(norm_val, 2)).font = _body_font()
        ws.cell(r, hdr_c + 2).border = _border
        ws.cell(r, hdr_c + 2).alignment = _center

    # Bar chart for counts
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Grain Size Distribution"
    bar_chart.y_axis.title = "Number of Grains"
    bar_chart.x_axis.title = f"Equivalent Diameter ({diam_unit})"
    bar_chart.style = 10
    bar_chart.width = 20
    bar_chart.height = 13
    bar_chart.legend = None  # No legend

    # Bar data
    bar_data = Reference(ws, min_col=hdr_c + 1,
                         min_row=data_row, max_row=data_row + n_bins)
    bar_chart.add_data(bar_data, titles_from_data=True)

    # Category labels (bin ranges under bars)
    cat_ref = Reference(ws, min_col=hdr_c,
                        min_row=data_row + 1, max_row=data_row + n_bins)
    bar_chart.set_categories(cat_ref)

    # Style bars — blue fill, thin border
    if bar_chart.series:
        s = bar_chart.series[0]
        s.graphicalProperties.solidFill = "4472C4"
        s.graphicalProperties.line.solidFill = "2B4570"
        s.graphicalProperties.line.width = 8000  # thin border in EMU

    # X-axis label rotation for readability
    bar_chart.x_axis.tickLblPos = "low"
    bar_chart.x_axis.txPr = None  # let Excel auto-rotate if needed

    # Line chart for normal curve
    line_chart = LineChart()
    line_data = Reference(ws, min_col=hdr_c + 2,
                          min_row=data_row, max_row=data_row + n_bins)
    line_chart.add_data(line_data, titles_from_data=True)
    line_chart.legend = None

    # Style the curve — thick magenta/pink line like the reference image
    if line_chart.series:
        s = line_chart.series[0]
        s.graphicalProperties.solidFill = None
        s.graphicalProperties.line.solidFill = "C0399F"
        s.graphicalProperties.line.width = 28000  # ~2pt
        s.smooth = True

    # Combine: overlay line on bar chart
    bar_chart += line_chart

    chart_col = get_column_letter(hdr_c + 4)
    ws.add_chart(bar_chart, f"{chart_col}{data_row}")

    _set_col_w(ws, hdr_c, 18)
    _set_col_w(ws, hdr_c + 1, 10)
    _set_col_w(ws, hdr_c + 2, 12)

    return data_row + n_bins + 2


# ==================================================================
# Summary table
# ==================================================================

def _write_summary_table(ws, result, image_path, start_row, start_col=1):
    area_unit, area_mult, diam_unit, diam_mult = _choose_unit(result.px_per_um)
    has_cal = result.has_calibration
    r = start_row

    def stat_row(label, value, unit="", alt=False):
        nonlocal r
        fill = _alt_fill(alt)
        c = ws.cell(r, start_col, label)
        c.font = _body_font(bold=True); c.fill = fill; c.border = _border
        c.alignment = _left_indent
        c = ws.cell(r, start_col + 1, value)
        c.font = _body_font(); c.fill = fill; c.border = _border; c.alignment = _center
        c = ws.cell(r, start_col + 2, unit)
        c.font = _body_font(color="666666"); c.fill = fill; c.border = _border
        c.alignment = _center
        r += 1

    for col in range(start_col, start_col + 3):
        c = ws.cell(r, col)
        c.fill = _hdr_fill(C_MID_BLUE); c.font = _hdr_font(10); c.alignment = _center
    ws.cell(r, start_col, "Statistic")
    ws.cell(r, start_col + 1, "Value")
    ws.cell(r, start_col + 2, "Unit")
    r += 1

    stat_row("Image", os.path.basename(image_path) if image_path else "N/A")
    stat_row("Total Grains", result.grain_count, "grains", True)

    if has_cal:
        stat_row("Mean Area", _fmt_num(result.mean_area_um2 * area_mult), area_unit)
        stat_row("Std Dev Area", _fmt_num(result.std_area_um2 * area_mult), area_unit, True)
        stat_row("Mean Diameter", _fmt_num(result.mean_diameter_um * diam_mult), diam_unit)
        stat_row("Std Dev Diameter", _fmt_num(result.std_diameter_um * diam_mult), diam_unit, True)
    else:
        areas_px = [g.area_px for g in result.grains] if result.grains else []
        if areas_px:
            stat_row("Mean Area", f"{np.mean(areas_px):.1f}", "px²")
            stat_row("Std Dev Area", f"{np.std(areas_px):.1f}", "px²", True)

    stat_row("Mean Circularity", f"{result.mean_circularity:.4f}", "(0–1)")
    stat_row("Mean Aspect Ratio", f"{result.mean_aspect_ratio:.4f}", "(1=equiaxed)", True)

    if has_cal and result.total_analyzed_area_um2 > 0:
        stat_row("Grain Coverage", f"{result.grain_coverage_pct:.2f}", "%")

    return r


# ==================================================================
# Main export functions
# ==================================================================

def export_to_excel(result, image_path, output_path, params=None, image_bgr=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_overview_sheet(wb, [(result, image_path, image_bgr)])
    _write_image_summary_sheet(wb, result, image_path, image_bgr, 1)
    _write_image_data_sheet(wb, result, image_path, 1)
    wb.save(output_path)
    return output_path


def export_multi_to_excel(analysed_tabs, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_overview_sheet(wb, analysed_tabs)
    for idx, (result, image_path, image_bgr) in enumerate(analysed_tabs, 1):
        _write_image_summary_sheet(wb, result, image_path, image_bgr, idx)
        _write_image_data_sheet(wb, result, image_path, idx)
    wb.save(output_path)
    return output_path


# ==================================================================
# Sheet writers
# ==================================================================

def _write_overview_sheet(wb, analysed_tabs):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "SEM Grain Analysis Report"
    c.font = Font(name="Arial", size=18, bold=True, color=C_WHITE)
    c.fill = _hdr_fill(C_DARK_BLUE)
    c.alignment = _center
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}    |    {len(analysed_tabs)} image(s)"
    c.font = Font(name="Arial", size=10, color=C_WHITE)
    c.fill = _hdr_fill(C_MID_BLUE)
    c.alignment = _center

    all_grains = []
    px_per_um = 0.0
    for result, path, img in analysed_tabs:
        all_grains.extend(result.grains)
        if result.has_calibration:
            px_per_um = result.px_per_um

    row = 4
    if all_grains:
        combined = type(analysed_tabs[0][0])()
        combined.grains = all_grains
        combined.grain_count = len(all_grains)
        combined.px_per_um = px_per_um
        combined.has_calibration = px_per_um > 0

        if combined.has_calibration:
            areas = np.array([g.area_um2 for g in all_grains])
            diams = np.array([g.equivalent_diameter_um for g in all_grains])
            combined.mean_area_um2 = float(np.mean(areas))
            combined.std_area_um2 = float(np.std(areas))
            combined.mean_diameter_um = float(np.mean(diams))
            combined.std_diameter_um = float(np.std(diams))
            total_area = sum(r.total_analyzed_area_um2
                             for r, p, i in analysed_tabs
                             if r.total_analyzed_area_um2 > 0)
            combined.total_analyzed_area_um2 = total_area
            if total_area > 0:
                combined.grain_coverage_pct = float(np.sum(areas)) / total_area * 100.0
            else:
                combined.grain_coverage_pct = 0.0
        else:
            combined.grain_coverage_pct = 0.0

        combined.mean_circularity = float(np.mean([g.circularity for g in all_grains]))
        combined.mean_aspect_ratio = float(np.mean([g.aspect_ratio for g in all_grains]))

        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row, 1, "Combined Summary (All Images)")
        c.font = _hdr_font(12); c.fill = _hdr_fill(C_DARK_BLUE); c.alignment = _center
        row += 1
        row = _write_summary_table(ws, combined, "All Images", row)
        row += 1

        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row, 1, "Grain Size Distribution (All Images)")
        c.font = _hdr_font(12); c.fill = _hdr_fill(C_DARK_BLUE); c.alignment = _center
        row += 1
        row = _write_histogram_chart(ws, all_grains, px_per_um, row)

    _set_col_w(ws, 1, 28); _set_col_w(ws, 2, 18); _set_col_w(ws, 3, 16)


def _write_image_summary_sheet(wb, result, image_path, image_bgr, idx):
    base = os.path.splitext(os.path.basename(image_path))[0][:18]
    ws = wb.create_sheet(f"{base}-Summary"[:31])
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"Image {idx}: {os.path.basename(image_path)}"
    c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
    c.fill = _hdr_fill(C_DARK_BLUE); c.alignment = _center
    ws.row_dimensions[1].height = 30

    row = 3

    if image_bgr is not None:
        try:
            tmp_orig = _save_image_temp(image_bgr, max_width=380)
            img_orig = XlImage(tmp_orig)
            ws.add_image(img_orig, f"A{row}")

            if result.overlay_image is not None:
                tmp_overlay = _save_image_temp(result.overlay_image, max_width=380)
                img_overlay = XlImage(tmp_overlay)
                ws.add_image(img_overlay, f"F{row}")

            h_img = image_bgr.shape[0]
            w_img = image_bgr.shape[1]
            scale = min(380 / w_img, 1.0)
            img_rows = max(15, int(h_img * scale / 15))
            row += img_rows
        except Exception:
            row += 2

    row += 1

    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row, 1, "Summary Statistics")
    c.font = _hdr_font(11); c.fill = _hdr_fill(C_DARK_BLUE); c.alignment = _center
    row += 1
    row = _write_summary_table(ws, result, image_path, row)
    row += 1

    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row, 1, "Grain Size Distribution")
    c.font = _hdr_font(11); c.fill = _hdr_fill(C_DARK_BLUE); c.alignment = _center
    row += 1
    if result.grains:
        _write_histogram_chart(ws, result.grains, result.px_per_um, row)

    _set_col_w(ws, 1, 28); _set_col_w(ws, 2, 18); _set_col_w(ws, 3, 16)


def _write_image_data_sheet(wb, result, image_path, idx):
    base = os.path.splitext(os.path.basename(image_path))[0][:18]
    ws = wb.create_sheet(f"{base}-Data"[:31])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    has_cal = result.has_calibration
    area_unit, area_mult, diam_unit, diam_mult = _choose_unit(result.px_per_um)

    if has_cal:
        headers = [
            "Grain ID", f"Area ({area_unit})", f"Equiv. Diameter ({diam_unit})",
            f"Major Axis ({diam_unit})", f"Minor Axis ({diam_unit})",
            f"Perimeter ({diam_unit})", "Circularity", "Aspect Ratio",
            "Eccentricity", "Centroid X (px)", "Centroid Y (px)",
        ]
    else:
        headers = [
            "Grain ID", "Area (px²)", "Equiv. Diameter (px)",
            "Perimeter (px)", "Circularity", "Aspect Ratio",
            "Eccentricity", "Centroid X (px)", "Centroid Y (px)",
        ]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws["A1"]
    c.value = f"Grain Data — {os.path.basename(image_path)}"
    c.font = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    c.fill = _hdr_fill(C_DARK_BLUE); c.alignment = _center

    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font = _hdr_font(10); c.fill = _hdr_fill(C_MID_BLUE)
        c.alignment = _center; c.border = _border

    for r_idx, grain in enumerate(result.grains, 3):
        alt = (r_idx % 2 == 0)
        fill = _alt_fill(alt)

        if has_cal:
            row_data = [
                grain.grain_id,
                round(grain.area_um2 * area_mult, 4),
                round(grain.equivalent_diameter_um * diam_mult, 4),
                round(grain.major_axis_um * diam_mult, 4),
                round(grain.minor_axis_um * diam_mult, 4),
                round(grain.perimeter_um * diam_mult, 4),
                round(grain.circularity, 4), round(grain.aspect_ratio, 4),
                round(grain.eccentricity, 4),
                round(grain.centroid_x, 1), round(grain.centroid_y, 1),
            ]
        else:
            row_data = [
                grain.grain_id, round(grain.area_px, 1),
                round(grain.equivalent_diameter_px, 2),
                round(grain.perimeter_px, 2),
                round(grain.circularity, 4), round(grain.aspect_ratio, 4),
                round(grain.eccentricity, 4),
                round(grain.centroid_x, 1), round(grain.centroid_y, 1),
            ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(r_idx, col, val)
            c.font = _body_font(); c.fill = fill; c.border = _border
            c.alignment = _center

    widths_cal = [10, 16, 18, 16, 16, 16, 14, 14, 14, 14, 14]
    widths_nocal = [10, 14, 18, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths_cal if has_cal else widths_nocal, 1):
        _set_col_w(ws, i, w)

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{2 + len(result.grains)}"
